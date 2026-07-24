from pathlib import Path
from hashlib import sha256
import re
import codecs
import json
import csv
import os
from io import StringIO, BytesIO
from typing import Any, Iterable
import pymupdf4llm
import base64

from dotenv import load_dotenv
from openai import OpenAI

from qdrant_client.models import (
    Distance,
    VectorParams,
    PointStruct,
    SparseVectorParams,
    SparseVector,
    PayloadSchemaType,
)

try:
    from .load_documents import config as cf
except ImportError:
    import config as cf

# ─── Parsers optionnels pour formats enrichis ───────────────────────────────
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from rdflib import Graph, Namespace, URIRef, Literal
    from rdflib.namespace import RDF, RDFS, SKOS
except ImportError:
    Graph = Namespace = URIRef = Literal = RDF = RDFS = SKOS = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import yaml
except ImportError:
    yaml = None

try:
    from striprtf.striprtf import rtf_to_text
except ImportError:
    rtf_to_text = None


# ─── Chargement du .env à la racine du projet ───────────────────────────────
_ROOT_ENV = Path(__file__).resolve().parent
while _ROOT_ENV.parent != _ROOT_ENV:
    if (_ROOT_ENV / ".env").exists():
        break
    _ROOT_ENV = _ROOT_ENV.parent
load_dotenv(_ROOT_ENV / ".env")

# ─── Client LLM (API OpenAI-compatible) ──────────────────────────────────
_LLM_API_KEY = os.getenv("LLM_API_KEY", "")
_URL_API = os.getenv("URL_LLM_API", "")
_llm_client = OpenAI(
    base_url=_URL_API,
    api_key=_LLM_API_KEY,
)
_LLM_MODEL = os.getenv("LLM_MODEL", "")
_SUMMARY_MAX_INPUT_CHARS = 500_000

client = cf.client
model = cf.model
COLLECTION = cf.COLLECTION
BATCH_SIZE = cf.BATCH_SIZE


XML_DECL_RE = re.compile(br'<\?xml[^>]+encoding\s*=\s*["\']([^"\']+)', re.IGNORECASE)
TEXT_BOMS = (
    codecs.BOM_UTF8,
    codecs.BOM_UTF16_LE,
    codecs.BOM_UTF16_BE,
    codecs.BOM_UTF32_LE,
    codecs.BOM_UTF32_BE,
)
TEXT_WHITELIST = set(b"\t\n\r")
PRINTABLE_ASCII = set(range(32, 127))

DENSE_VECTOR_NAME = "dense"
SPARSE_VECTOR_NAME = "sparse"

CHUNK_SIZE = int(
    getattr(cf, "config", {}).get("chunking", {}).get("chunk_size", 4000)
    if hasattr(cf, "config")
    else 4000
)

CHUNK_OVERLAP = int(
    getattr(cf, "config", {}).get("chunking", {}).get("chunk_overlap", 400)
    if hasattr(cf, "config")
    else 400
)

_SUMMARY_SYSTEM_PROMPT = (
    "Tu es un assistant expert en résumé de documents techniques et réglementaires. "
    "Produis un résumé factuel, neutre et très concis du document fourni. "
    "Le résumé doit :\n"
    "- Présenter l'objet principal du document en une phrase.\n"
    "- Optionnellement lister 1 à 2 points clés.\n"
    "- Ne rien inventer ni extrapoler au-delà du texte fourni.\n"
    "- Être rédigé en français.\n"
    "- Avoir une longueur totale comprise entre 200 et 500 caractères MAXIMUM."
)


# ─── Génération du résumé ────────────────────────────────────────────────────

def generate_doc_summary(filename: str, full_text: str) -> str:
    """
    Appelle un LLM (via API OpenAI-compatible) pour produire un résumé du document.
    Tronque à _SUMMARY_MAX_INPUT_CHARS si nécessaire.
    Retourne une chaîne vide en cas d'échec (non bloquant pour l'indexation).
    """
    if not full_text.strip():
        return ""

    truncated = False
    text_input = full_text
    if len(full_text) > _SUMMARY_MAX_INPUT_CHARS:
        text_input = full_text[:_SUMMARY_MAX_INPUT_CHARS]
        truncated = True
        print(
            f"[~] '{filename}' truncated for summary: "
            f"{len(full_text)} → {_SUMMARY_MAX_INPUT_CHARS} chars"
        )

    user_message = (
        f"Voici le contenu du document '{filename}'"
        + (" (tronqué, suite non disponible)" if truncated else "")
        + f" :\n\n{text_input}\n\nRésume ce document selon les instructions."
    )

    try:
        response = _llm_client.chat.completions.create(
            model=_LLM_MODEL,
            messages=[
                {"role": "system", "content": _SUMMARY_SYSTEM_PROMPT},
                {"role": "user", "content": user_message},
            ],
            temperature=0.0,
        )
        summary = response.choices[0].message.content or ""
        return summary.strip()
    except Exception as e:
        print(f"[!] Summary generation failed for '{filename}': {e}")
        return ""


# ─── Détection des capacités du modèle ──────────────────────────────────────

def detect_model_capabilities(model) -> dict[str, Any]:
    capabilities = {
        "has_dense": False,
        "has_sparse": False,
        "dense_dim": None,
    }

    test_text = "test"

    try:
        result = model.encode(
            [test_text],
            return_dense=True,
            return_sparse=True,
            return_colbert_vecs=False,
        )

        if isinstance(result, dict):
            dense_vecs = result.get("dense_vecs")
            lexical_weights = result.get("lexical_weights")

            if dense_vecs is not None:
                first_dense = dense_vecs[0] if len(dense_vecs) > 0 else None
                if first_dense is not None:
                    capabilities["has_dense"] = True
                    capabilities["dense_dim"] = len(first_dense)

            if lexical_weights is not None:
                capabilities["has_sparse"] = True

            print(
                f"Model capabilities detected: "
                f"dense={capabilities['has_dense']}, "
                f"sparse={capabilities['has_sparse']}, "
                f"dense_dim={capabilities['dense_dim']}"
            )
            return capabilities

    except Exception:
        pass

    try:
        dense = model.encode([test_text])
        first_dense = dense[0] if hasattr(dense, "__len__") else dense
        capabilities["has_dense"] = True
        capabilities["dense_dim"] = len(first_dense)
        print(
            f"Model capabilities detected: dense=True, sparse=False, "
            f"dense_dim={capabilities['dense_dim']}"
        )
        return capabilities
    except Exception as e:
        raise ValueError(f"Could not detect model capabilities: {e}")


# ─── Détection d'encodage ────────────────────────────────────────────────────

def detect_bom_encoding(raw: bytes) -> str | None:
    if raw.startswith(codecs.BOM_UTF8):
        return "utf-8-sig"
    if raw.startswith(codecs.BOM_UTF32_LE):
        return "utf-32-le"
    if raw.startswith(codecs.BOM_UTF32_BE):
        return "utf-32-be"
    if raw.startswith(codecs.BOM_UTF16_LE):
        return "utf-16-le"
    if raw.startswith(codecs.BOM_UTF16_BE):
        return "utf-16-be"
    return None


def detect_xml_decl_encoding(raw: bytes) -> str | None:
    head = raw[:2048]
    match = XML_DECL_RE.search(head)
    if match:
        try:
            return match.group(1).decode("ascii").lower()
        except Exception:
            return None
    return None


def is_probably_binary(filepath: Path, chunk_size: int = 4096) -> bool:
    raw = filepath.read_bytes()
    if not raw:
        return False

    sample = raw[:chunk_size]

    if any(sample.startswith(bom) for bom in TEXT_BOMS):
        return False

    if b"\x00" in sample:
        return True

    odd = 0
    for b in sample:
        if b in TEXT_WHITELIST or b in PRINTABLE_ASCII:
            continue
        if b >= 128:
            continue
        odd += 1

    return (odd / len(sample)) > 0.30


def guess_text_encodings(raw: bytes) -> list[str]:
    candidates = []

    bom_encoding = detect_bom_encoding(raw)
    if bom_encoding:
        candidates.append(bom_encoding)

    xml_decl_encoding = detect_xml_decl_encoding(raw)
    if xml_decl_encoding:
        candidates.append(xml_decl_encoding)

    candidates.extend(
        [
            "utf-8",
            "utf-8-sig",
            "utf-16",
            "utf-16-le",
            "utf-16-be",
            "utf-32",
            "utf-32-le",
            "utf-32-be",
            "cp1252",
            "latin-1",
        ]
    )

    deduped = []
    seen = set()
    for enc in candidates:
        if enc and not enc in seen:
            deduped.append(enc)
            seen.add(enc)

    return deduped


# ─── Lecture des fichiers ────────────────────────────────────────────────────

def read_text_document(filepath: Path) -> tuple[str, str]:
    raw = filepath.read_bytes()

    if is_probably_binary(filepath):
        raise ValueError(f"{filepath.name} appears to be a binary/non-text file")

    encodings_to_try = guess_text_encodings(raw)

    for encoding in encodings_to_try:
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
        except LookupError:
            continue

    raise ValueError(
        f"Unable to decode file {filepath.name}. "
        f"Tried encodings: {', '.join(encodings_to_try)}"
    )


# ─── Parsers spécialisés par type de document ───────────────────────────────

RE_ARTICLE_HEADER = re.compile(
    r"^(Article|Art\.?)\s+([LRD]?\d+[\d\-\.]*(?:\s*-\s*\d+)?)\.?\s*$",
    re.IGNORECASE | re.MULTILINE,
)
RE_ARTICLE_MENTION = re.compile(
    r"\b(?:article|art\.?)\s+([LRD]?\d+[\d\-\.]*(?:\s*-\s*\d+)?)\b",
    re.IGNORECASE,
)


def _normalize_article_ref(match_text: str) -> str:
    """Normalise une référence d'article pour matching."""
    return re.sub(r"\s+", "", match_text.upper()).replace("ART.", "").replace("ARTICLE", "")


def parse_legal_articles(text: str) -> list[tuple[str | None, str]]:
    """
    Découpe un texte juridique en articles.
    Retourne une liste de (numéro_article_ou_None, contenu_article).
    """
    parts = RE_ARTICLE_HEADER.split(text)
    if len(parts) <= 1:
        return [(None, text)]

    result = []
    current_article = None
    current_buffer = []

    i = 0
    while i < len(parts):
        if parts[i] is None:
            i += 1
            continue
        if re.match(r"^(article|art\.?)$", parts[i], re.IGNORECASE):
            # Flush previous
            if current_buffer:
                content = "\n".join(current_buffer).strip()
                if content:
                    result.append((current_article, content))
                current_buffer = []
            current_article = parts[i + 1].strip() if i + 1 < len(parts) else None
            i += 2
            continue
        current_buffer.append(parts[i])
        i += 1

    if current_buffer:
        content = "\n".join(current_buffer).strip()
        if content:
            result.append((current_article, content))

    return result


def extract_cited_articles(text: str) -> list[str]:
    """Extrait les mentions d'articles cités dans un texte."""
    refs = set()
    for m in RE_ARTICLE_MENTION.finditer(text):
        refs.add(_normalize_article_ref(m.group(1)))
    return sorted(refs)


def _detect_term_column(headers: list[str], rows: list[dict[str, str]]) -> str | None:
    """Détecte automatiquement la colonne contenant le terme/concept."""
    lower_headers = [h.strip().lower() for h in headers]

    term_candidates = [
        "terme", "term", "terme français", "terme anglais", "libellé", "label",
        "mot", "concept", "entrée", "intitulé", "name", "title", "mot-clé",
        "mot clef", "keyword", "expression", "forme", "vocable", "appellation",
        "dénomination", "nom", "titre", "acronyme", "sigle", "abréviation",
        "term_id", "id", "code", "identifiant",
    ]
    for term_candidate in term_candidates:
        if term_candidate in lower_headers:
            return headers[lower_headers.index(term_candidate)]

    # Heuristique sur le contenu : colonne avec les textes les plus courts en moyenne
    if rows:
        avg_lengths = []
        for h in headers:
            lengths = [len(row.get(h, "").strip()) for row in rows[:50] if row.get(h, "").strip()]
            if lengths:
                avg_lengths.append((h, sum(lengths) / len(lengths)))
        if avg_lengths:
            avg_lengths.sort(key=lambda x: x[1])
            # On prend la colonne la plus courte, à condition qu'elle soit significativement plus courte que la suivante
            if len(avg_lengths) >= 2 and avg_lengths[0][1] * 3 < avg_lengths[1][1]:
                return avg_lengths[0][0]
            # Sinon on prend la première colonne si elle a des valeurs courtes
            if avg_lengths[0][1] <= 100:
                return avg_lengths[0][0]

    # Fallback : première colonne non vide
    return headers[0] if headers else None


def _detect_definition_column(headers: list[str], rows: list[dict[str, str]]) -> str | None:
    """Détecte automatiquement la colonne contenant la définition."""
    lower_headers = [h.strip().lower() for h in headers]

    def_candidates = [
        "définition", "definition", "définitions", "definitions", "sens",
        "signification", "description", "explication", "note", "notes",
        "commentaire", "remarque", "exemple", "contexte", "usage", "domaine",
        "meaning", "sense", "gloss", "explanatory", "scope note",
    ]
    for def_candidate in def_candidates:
        if def_candidate in lower_headers:
            return headers[lower_headers.index(def_candidate)]

    # Heuristique sur le contenu : colonne avec les textes les plus longs en moyenne
    if rows:
        avg_lengths = []
        for h in headers:
            lengths = [len(row.get(h, "").strip()) for row in rows[:50] if row.get(h, "").strip()]
            if lengths:
                avg_lengths.append((h, sum(lengths) / len(lengths)))
        if avg_lengths:
            avg_lengths.sort(key=lambda x: x[1], reverse=True)
            if avg_lengths[0][1] >= 30:
                return avg_lengths[0][0]

    return None


def _has_header_row(raw_lines: list[str], delimiter: str) -> bool:
    """Détecte si la première ligne est un en-tête ou une donnée."""
    if len(raw_lines) < 2:
        return False
    first_line = raw_lines[0].strip()
    parts = first_line.split(delimiter)
    if len(parts) <= 1:
        return False

    # Heuristique 1 : mots-clés d'en-tête connus
    known_header_keywords = {
        "terme", "term", "définition", "definition", "libellé", "label",
        "description", "name", "concept", "titre", "title", "mot", "entrée",
        "explication", "domaine", "source", "note", "commentaire", "exemple",
        "id", "code", "identifiant", "acronyme", "sigle", "abréviation",
    }
    keyword_score = 0
    for p in parts:
        p_norm = p.strip().lower().rstrip("s")
        if p_norm in known_header_keywords:
            keyword_score += 1

    if keyword_score >= 1 and keyword_score >= len(parts) * 0.25:
        return True

    # Heuristique 2 : comparer ligne 1 et ligne 2
    # Une ligne d'en-tête est typiquement plus courte et a une structure différente de la ligne suivante
    second_line = raw_lines[1].strip()
    second_parts = second_line.split(delimiter)
    if len(second_parts) != len(parts):
        return False

    first_avg_len = sum(len(p.strip()) for p in parts) / len(parts)
    second_avg_len = sum(len(p.strip()) for p in second_parts) / len(parts)

    # L'en-tête est nettement plus court que la première ligne de données
    if second_avg_len > first_avg_len * 2 and first_avg_len <= 30:
        return True

    # Heuristique 3 : la première ligne ressemble à des noms courts, la deuxième à du contenu
    first_is_short = sum(1 for p in parts if len(p.strip().split()) <= 3) >= len(parts) * 0.7
    second_is_longer = second_avg_len > first_avg_len * 1.5

    if first_is_short and second_is_longer and first_avg_len <= 40:
        return True

    return False


def chunk_glossary_table(text: str, delimiter: str = ",") -> list[dict[str, Any]]:
    """
    Parse un fichier de glossaire tabulaire (CSV/TSV).
    Chaque ligne devient un chunk avec terme/définition.
    Fonctionne avec ou sans ligne d'en-tête.
    """
    raw_lines = [line for line in text.splitlines() if line.strip()]
    if not raw_lines:
        return []

    has_header = _has_header_row(raw_lines, delimiter)

    if has_header:
        reader = csv.DictReader(StringIO(text), delimiter=delimiter)
        rows = list(reader)
        headers = list(rows[0].keys()) if rows else []
    else:
        # Sans en-tête : on génère des noms de colonnes
        first_parts = raw_lines[0].split(delimiter)
        headers = [f"col_{i}" for i in range(len(first_parts))]
        rows = []
        for line in raw_lines:
            parts = line.split(delimiter)
            # Gérer le cas où il y aurait plus ou moins de colonnes
            row = {headers[i]: parts[i].strip() if i < len(parts) else "" for i in range(len(headers))}
            rows.append(row)

    if not rows:
        return []

    term_col = _detect_term_column(headers, rows)
    definition_col = _detect_definition_column(headers, rows)

    chunks = []
    for idx, row in enumerate(rows):
        if not any(v.strip() for v in row.values()):
            continue

        term = row.get(term_col, "").strip() if term_col else ""
        definition = row.get(definition_col, "").strip() if definition_col else ""

        lines = []
        for h in headers:
            v = row.get(h, "").strip()
            if v:
                lines.append(f"{h.strip()} : {v}")
        body = "\n".join(lines)

        # Si on n'a pas trouvé de terme, utiliser le premier champ non vide
        if not term and body:
            term = next((row.get(h, "").strip() for h in headers if row.get(h, "").strip()), "")

        chunks.append({
            "text": body,
            "doctype": "glossary_entry",
            "term": term,
            "definition": definition,
            "chunk_index": idx,
        })

    return chunks


def chunk_excel(filepath: Path) -> list[dict[str, Any]]:
    """Parse un fichier Excel de glossaire ligne par ligne."""
    if openpyxl is None:
        raise ImportError("openpyxl is required for .xlsx files")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    chunks = []
    idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue

        # Détecter si la première ligne est un en-tête
        first_values = [cell.value for cell in rows[0]]
        has_header = _has_header_row(
            [delimiter.join(str(v or "") for v in first_values)],
            delimiter=","
        )

        if has_header:
            header_row = rows[0]
            headers = []
            for cell in header_row:
                val = cell.value
                headers.append(str(val).strip() if val is not None else f"col_{len(headers)}")
            data_rows = rows[1:]
        else:
            headers = [f"col_{i}" for i in range(len(first_values))]
            data_rows = rows

        term_col = _detect_term_column(headers, [
            {h: str(cell.value or "").strip() for h, cell in zip(headers, row)} for row in data_rows[:50]
        ])
        definition_col = _detect_definition_column(headers, [
            {h: str(cell.value or "").strip() for h, cell in zip(headers, row)} for row in data_rows[:50]
        ])

        for raw_row in data_rows:
            values = [cell.value for cell in raw_row]
            if all(v is None or str(v).strip() == "" for v in values):
                continue

            row = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(values)}
            term = row.get(term_col, "") if term_col else ""
            definition = row.get(definition_col, "") if definition_col else ""

            lines = [f"{h} : {v}" for h, v in row.items() if v]
            body = "\n".join(lines)

            if not term and body:
                term = next((row.get(h, "").strip() for h in headers if row.get(h, "").strip()), "")

            chunks.append({
                "text": body,
                "doctype": "glossary_entry",
                "term": term,
                "definition": definition,
                "chunk_index": idx,
                "sheet": sheet_name,
            })
            idx += 1

    return chunks


def _rdf_text_value(graph, subject, predicate) -> str:
    if graph is None:
        return ""
    for obj in graph.objects(subject, predicate):
        if isinstance(obj, Literal):
            return str(obj)
    return ""


def _rdf_list_values(graph, subject, predicate) -> list[str]:
    if graph is None:
        return []
    out = []
    for obj in graph.objects(subject, predicate):
        if isinstance(obj, Literal):
            out.append(str(obj))
    return out


def _rdf_uri_list(graph, subject, predicate) -> list[str]:
    if graph is None:
        return []
    return [str(obj) for obj in graph.objects(subject, predicate)]


def _rdf_expand_namespace(graph, uri: str) -> str:
    """Essaye de convertir une URI compacte en URI complète si possible."""
    if isinstance(uri, URIRef):
        return str(uri)
    return str(uri)


def chunk_rdf(filepath: Path, text: str, base_document_id: str) -> list[dict[str, Any]]:
    """
    Parse un fichier RDF et retourne un chunk par concept/ressource.
    Accepte SKOS, OWL, RDFS et d'autres vocabulaires.
    """
    if Graph is None:
        raise ImportError("rdflib is required for RDF files")

    g = Graph()

    # Essayer plusieurs formats de sérialisation
    parsed = False
    for fmt in ("turtle", "xml", "json-ld", "n3", "nt", "trig"):
        try:
            g.parse(data=text, format=fmt)
            parsed = True
            break
        except Exception:
            continue

    if not parsed:
        try:
            g.parse(data=text)
            parsed = True
        except Exception as e:
            raise ValueError(f"Could not parse RDF file {filepath.name}: {e}")

    # Namespaces connus pour les liens hiérarchiques
    RDFS = Namespace("http://www.w3.org/2000/01/rdf-schema#")
    OWL = Namespace("http://www.w3.org/2002/07/owl#")
    DC = Namespace("http://purl.org/dc/elements/1.1/")
    DCTERMS = Namespace("http://purl.org/dc/terms/")
    SCHEMA = Namespace("http://schema.org/")

    # Prédicats de libellés (ordre de priorité)
    label_predicates = [
        SKOS.prefLabel,
        RDFS.label,
        DCTERMS.title,
        DC.title,
        SCHEMA.name,
        SKOS.altLabel,
    ]

    # Prédicats de définition (ordre de priorité)
    definition_predicates = [
        SKOS.definition,
        RDFS.comment,
        DCTERMS.description,
        DC.description,
        SCHEMA.description,
        SKOS.scopeNote,
        SKOS.note,
        RDFS.isDefinedBy,
    ]

    # Prédicats de synonymes
    synonym_predicates = [
        SKOS.altLabel,
        SKOS.hiddenLabel,
        RDFS.label,
        SCHEMA.alternateName,
    ]

    # Prédicats de liens hiérarchiques/associatifs
    broader_predicates = [
        SKOS.broader,
        SKOS.broaderTransitive,
        RDFS.subClassOf,
        DCTERMS.isPartOf,
        SCHEMA.isPartOf,
    ]
    narrower_predicates = [
        SKOS.narrower,
        SKOS.narrowerTransitive,
    ]
    related_predicates = [
        SKOS.related,
        SKOS.relatedMatch,
        SKOS.closeMatch,
        SKOS.exactMatch,
        SKOS.broadMatch,
        SKOS.narrowMatch,
        RDFS.seeAlso,
        OWL.sameAs,
        SCHEMA.sameAs,
    ]

    # Déterminer quelles ressources sont des "concepts" à indexer
    # On indexe toute ressource qui a au moins un libellé, une définition, ou un type connu
    concept_types = {
        SKOS.Concept,
        SKOS.Collection,
        RDFS.Resource,
        RDFS.Class,
        OWL.Class,
        OWL.NamedIndividual,
        SCHEMA.DefinedTerm,
    }

    candidates = set()
    for s in g.subjects():
        if isinstance(s, URIRef) or isinstance(s, str):
            candidates.add(s)

    # Filtrer : garder les ressources avec au moins un libellé ou une définition, ou un type concept
    resources_to_index = []
    for r in candidates:
        has_type = False
        for t in g.objects(r, RDF.type):
            if t in concept_types:
                has_type = True
                break

        has_label = any(_rdf_list_values(g, r, pred) for pred in label_predicates)
        has_definition = any(_rdf_list_values(g, r, pred) for pred in definition_predicates)

        if has_type or has_label or has_definition:
            resources_to_index.append(r)

    # Supprimer les ressources qui sont seulement des blank nodes or namespaces
    resources_to_index = [r for r in resources_to_index if isinstance(r, URIRef)]

    # Si aucune ressource avec type/libellé/définition, indexer toutes les URI
    if not resources_to_index:
        resources_to_index = [r for r in candidates if isinstance(r, URIRef)]

    chunks = []
    for idx, concept in enumerate(resources_to_index):
        # Libellé principal : prendre le premier trouvé
        pref = ""
        for pred in label_predicates:
            labels = _rdf_list_values(g, concept, pred)
            if labels:
                pref = labels[0]
                break

        # Définition principale
        definition = ""
        for pred in definition_predicates:
            defs = _rdf_list_values(g, concept, pred)
            if defs:
                definition = defs[0]
                break

        # Synonymes / labels alternatifs
        synonyms = []
        seen_synonyms = set()
        for pred in synonym_predicates:
            for val in _rdf_list_values(g, concept, pred):
                if val != pref and val not in seen_synonyms:
                    seen_synonyms.add(val)
                    synonyms.append(val)

        # Notes et autres descriptions
        notes = []
        for pred in [SKOS.note, SKOS.scopeNote, SKOS.example, SKOS.historyNote, SKOS.editorialNote, SKOS.changeNote]:
            notes.extend(_rdf_list_values(g, concept, pred))

        # Liens
        broaders = []
        seen_links = set()
        for pred in broader_predicates:
            for uri in _rdf_uri_list(g, concept, pred):
                if uri not in seen_links:
                    seen_links.add(uri)
                    broaders.append(uri)

        narrowers = []
        seen_links = set()
        for pred in narrower_predicates:
            for uri in _rdf_uri_list(g, concept, pred):
                if uri not in seen_links:
                    seen_links.add(uri)
                    narrowers.append(uri)

        relateds = []
        seen_links = set()
        for pred in related_predicates:
            for uri in _rdf_uri_list(g, concept, pred):
                if uri not in seen_links:
                    seen_links.add(uri)
                    relateds.append(uri)

        # Types
        types = _rdf_uri_list(g, concept, RDF.type)

        # Autres propriétés littérales intéressantes
        extra_lines = []
        interesting_preds = set()
        for p, o in g.predicate_objects(concept):
            if isinstance(o, Literal):
                pred_str = _rdf_expand_namespace(g, p)
                if pred_str not in interesting_preds:
                    interesting_preds.add(pred_str)
                    vals = _rdf_list_values(g, concept, p)
                    if vals and p not in label_predicates and p not in definition_predicates and p not in synonym_predicates:
                        short_name = pred_str.split("/")[-1].split("#")[-1]
                        extra_lines.append(f"{short_name} : {' | '.join(vals)}")

        lines = [f"Concept : {pref}"] if pref else [f"Concept URI : {concept}"]
        if types:
            lines.append(f"Types : {', '.join(t.split('/')[-1].split('#')[-1] for t in types)}")
        if synonyms:
            lines.append(f"Synonymes : {', '.join(synonyms)}")
        if definition:
            lines.append(f"Définition : {definition}")
        if notes:
            lines.append(f"Notes : {' | '.join(notes)}")
        if broaders:
            lines.append(f"Concepts plus généraux : {', '.join(broaders)}")
        if narrowers:
            lines.append(f"Concepts plus spécifiques : {', '.join(narrowers)}")
        if relateds:
            lines.append(f"Concepts liés : {', '.join(relateds)}")
        lines.extend(extra_lines)

        body = "\n".join(lines)
        if not body.strip():
            continue

        chunks.append({
            "text": body,
            "doctype": "rdf_concept",
            "concept_uri": str(concept),
            "term": pref,
            "definition": definition,
            "synonyms": synonyms,
            "broader": broaders,
            "narrower": narrowers,
            "related": relateds,
            "rdf_types": types,
            "chunk_index": idx,
        })

    return chunks


def chunk_json_structured(text: str) -> list[dict[str, Any]]:
    """
    Parse un JSON structuré (liste de concepts, glossaire JSON-LD-like).
    Retourne un chunk par entrée si possible.
    """
    try:
        data = json.loads(text)
    except Exception:
        return []

    chunks = []
    if isinstance(data, list):
        for idx, item in enumerate(data):
            if isinstance(item, dict):
                body = _format_dict(item)
                chunks.append({
                    "text": body,
                    "doctype": "json_entry",
                    "term": _first_value(item, ["terme", "term", "label", "name", "title", "prefLabel", "concept"]),
                    "definition": _first_value(item, ["définition", "definition", "description", "definitionText", "meaning"]),
                    "chunk_index": idx,
                })
    elif isinstance(data, dict):
        # Essayer de trouver une liste de concepts
        for key, value in data.items():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                for idx, item in enumerate(value):
                    body = _format_dict(item)
                    chunks.append({
                        "text": body,
                        "doctype": "json_entry",
                        "term": _first_value(item, ["terme", "term", "label", "name", "title", "prefLabel", "concept"]),
                        "definition": _first_value(item, ["définition", "definition", "description", "definitionText", "meaning"]),
                        "chunk_index": idx,
                        "parent_key": key,
                    })
                break
        else:
            # Sinon un chunk pour tout le document
            body = _format_dict(data)
            chunks.append({
                "text": body,
                "doctype": "json_entry",
                "term": _first_value(data, ["title", "name", "label"]),
                "chunk_index": 0,
            })

    return chunks


def _format_dict(item: dict) -> str:
    lines = []
    for k, v in sorted(item.items()) if isinstance(item, dict) else []:
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False)
        lines.append(f"{k} : {v}")
    return "\n".join(lines)


def _first_value(item: dict, keys: Iterable[str]) -> str:
    for k in keys:
        if k in item and item[k]:
            return str(item[k]).strip()
    return ""


def chunk_yaml_structured(text: str) -> list[dict[str, Any]]:
    """Parse un YAML structuré et chunk par entrée de premier niveau."""
    if yaml is None:
        return []
    try:
        data = yaml.safe_load(text)
    except Exception:
        return []
    return chunk_json_structured(json.dumps(data, ensure_ascii=False))


def _html_text(elem) -> str:
    """Extract clean text from a BeautifulSoup element."""
    return elem.get_text(" ", strip=True)


def chunk_eurlex_html(soup) -> list[dict[str, Any]]:
    """
    Détecte et découpe un document HTML Eur-Lex / Journal officiel de l'UE
    par chapitres et articles, en utilisant les classes CSS officielles.
    Les gros blocs (préambule, articles longs) sont subdivisés.
    """
    chunks = []
    current_chapter = ""
    current_section = ""
    current_article = ""
    current_body: list[str] = []
    idx = 0

    def _emit(body_parts: list[str], extra_title: str = "", article_override: str = "") -> None:
        nonlocal idx
        if not body_parts:
            return
        body = "\n".join(body_parts).strip()
        if not body:
            return
        title_parts = [p for p in [current_chapter, current_section, current_article, extra_title] if p]
        title = " — ".join(title_parts)
        text_chunk = f"{title}\n\n{body}".strip() if title else body
        chunks.append({
            "text": text_chunk,
            "doctype": "eurlex_article",
            "section_title": title,
            "article": article_override or current_article,
            "chunk_index": idx,
        })
        idx += 1

    def flush(subdivide: bool = True) -> None:
        nonlocal current_body
        if not current_body:
            current_body = []
            return
        title_parts = [p for p in [current_chapter, current_section, current_article] if p]
        title = " — ".join(title_parts)
        full_text = "\n".join(current_body).strip()
        current_body = []

        if not subdivide or len(full_text) <= CHUNK_SIZE:
            _emit([full_text])
            return

        # Subdivide large blocks while keeping the article/chapter context in each piece
        sub_chunks = split_text_uniformly(full_text, CHUNK_SIZE, CHUNK_OVERLAP)
        for sub in sub_chunks:
            _emit([sub])

    def flush_preamble(preamble_body: list[str]) -> None:
        nonlocal idx
        if not preamble_body:
            return
        body = "\n".join(preamble_body).strip()
        if not body:
            return
        if len(body) <= CHUNK_SIZE:
            _emit([body], extra_title="Préambule")
            return
        sub_chunks = split_text_uniformly(body, CHUNK_SIZE, CHUNK_OVERLAP)
        for sub in sub_chunks:
            _emit([sub], extra_title="Préambule")

    # Find the main content container; fallback to body
    main = soup.find("main") or soup.find("div", class_="eli-container") or soup.body or soup
    if main is None:
        return []

    preamble_body: list[str] = []
    in_preamble = True
    doc_title_parts: list[str] = []

    for elem in main.find_all(["p", "div"]):
        classes = elem.get("class", []) or []
        txt = _html_text(elem)
        if not txt:
            continue

        # Document main title (before preamble): collect consecutive oj-doc-ti lines
        if "oj-doc-ti" in classes:
            if not doc_title_parts:
                flush_preamble(preamble_body)
                preamble_body = []
                current_article = ""
                current_section = ""
                current_chapter = ""
                in_preamble = False
            doc_title_parts.append(txt)
            continue

        # Flush collected title once we hit the real preamble text or structure
        if doc_title_parts:
            _emit(["\n".join(doc_title_parts)], extra_title="Titre")
            doc_title_parts = []

        # We are still in preamble before any chapter/article
        if in_preamble and "oj-ti-section-1" not in classes and "oj-ti-section-2" not in classes and "oj-ti-art" not in classes:
            preamble_body.append(txt)
            continue

        # Chapter heading
        if "oj-ti-section-1" in classes:
            flush_preamble(preamble_body)
            preamble_body = []
            in_preamble = False
            flush()
            current_chapter = txt
            current_section = ""
            current_article = ""
            continue

        # Section heading
        if "oj-ti-section-2" in classes:
            flush()
            current_section = txt
            current_article = ""
            continue

        # Article heading
        if "oj-ti-art" in classes:
            flush()
            current_article = txt
            continue

        # Ordinary paragraph
        if in_preamble:
            preamble_body.append(txt)
        else:
            current_body.append(txt)

    if doc_title_parts:
        _emit(["\n".join(doc_title_parts)], extra_title="Titre")
    flush_preamble(preamble_body)
    flush()
    return chunks


def _clean_html_to_text(text: str) -> str:
    """Strip HTML tags and scripts, keeping paragraph structure."""
    if BeautifulSoup is None:
        return text
    soup = BeautifulSoup(text, "html.parser")
    for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
        tag.decompose()
    # Replace block elements with newlines to preserve some structure
    for tag in soup.find_all(["p", "div", "li", "br", "h1", "h2", "h3", "h4", "h5", "h6"]):
        tag.append("\n")
    return soup.get_text("\n", strip=True)


def chunk_html_glossary(text: str) -> list[dict[str, Any]]:
    """
    Extrait un glossaire web structuré : liste de termes/définitions.
    Fallback : documents Eur-Lex, puis chunk par titres, puis texte structuré.
    """
    if BeautifulSoup is None:
        return []

    soup = BeautifulSoup(text, "html.parser")
    # Nettoyage
    for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
        tag.decompose()

    chunks = []

    # Stratégie 1 : définitions HTML (dt/dd)
    dts = soup.find_all("dt")
    if dts:
        for idx, dt in enumerate(dts):
            term = _html_text(dt)
            dd = dt.find_next("dd")
            definition = _html_text(dd) if dd else ""
            body = f"Terme : {term}\nDéfinition : {definition}".strip()
            chunks.append({
                "text": body,
                "doctype": "glossary_entry",
                "term": term,
                "definition": definition,
                "chunk_index": idx,
            })
        return chunks

    # Stratégie 2 : documents Eur-Lex / Journal officiel de l'UE
    eurlex_chunks = chunk_eurlex_html(soup)
    if eurlex_chunks:
        return eurlex_chunks

    # Stratégie 3 : titres h2/h3 suivis de paragraphes
    sections = []
    current_title = ""
    current_body = []

    for elem in soup.find_all(["h1", "h2", "h3", "h4", "p", "li"]):
        if elem.name in ("h1", "h2", "h3", "h4"):
            if current_body:
                sections.append((current_title, "\n".join(current_body)))
            current_title = _html_text(elem)
            current_body = []
        else:
            txt = _html_text(elem)
            if txt:
                current_body.append(txt)
    if current_body:
        sections.append((current_title, "\n".join(current_body)))

    if sections:
        for idx, (title, body) in enumerate(sections):
            text_chunk = f"{title}\n\n{body}".strip() if title else body
            if text_chunk:
                chunks.append({
                    "text": text_chunk,
                    "doctype": "html_section",
                    "section_title": title,
                    "chunk_index": idx,
                })
        return chunks

    # Stratégie 4 : fallback texte structuré sur HTML nettoyé
    clean_text = _clean_html_to_text(str(soup))
    return chunk_text_structured(clean_text)


def chunk_pdf_by_sections(text: str) -> list[dict[str, Any]]:
    """Chunk un PDF converti en Markdown par titres de sections."""
    lines = text.splitlines()
    chunks = []
    current_title = ""
    current_level = 0
    current_body = []
    idx = 0

    title_re = re.compile(r"^(#{1,4})\s+(.+)$")

    for line in lines:
        m = title_re.match(line.strip())
        if m:
            if current_body:
                body = "\n".join(current_body).strip()
                if body:
                    chunks.append({
                        "text": f"{current_title}\n\n{body}".strip() if current_title else body,
                        "doctype": "pdf_section",
                        "section_title": current_title,
                        "section_level": current_level,
                        "chunk_index": idx,
                    })
                    idx += 1
            current_level = len(m.group(1))
            current_title = m.group(2).strip()
            current_body = []
        else:
            current_body.append(line)

    if current_body:
        body = "\n".join(current_body).strip()
        if body:
            chunks.append({
                "text": f"{current_title}\n\n{body}".strip() if current_title else body,
                "doctype": "pdf_section",
                "section_title": current_title,
                "section_level": current_level,
                "chunk_index": idx,
            })

    return chunks if chunks else [{"text": text, "doctype": "pdf_section", "chunk_index": 0}]


def chunk_text_structured(text: str) -> list[dict[str, Any]]:
    """
    Détecte la structure d'un texte brut ou Markdown :
    - articles de loi
    - sections par titres Markdown
    - fallback uniforme
    """
    # 1. Articles de loi
    articles = parse_legal_articles(text)
    if len(articles) > 1 or (len(articles) == 1 and articles[0][0] is not None):
        chunks = []
        for idx, (article, content) in enumerate(articles):
            cited = extract_cited_articles(content)
            title = f"Article {article}" if article else "Texte sans article"
            chunks.append({
                "text": f"{title}\n\n{content}".strip(),
                "doctype": "legal_article",
                "article": article,
                "cited_articles": cited,
                "chunk_index": idx,
            })
        return chunks

    # 2. Sections Markdown
    md_chunks = chunk_pdf_by_sections(text)
    if len(md_chunks) > 1:
        for c in md_chunks:
            c["doctype"] = "text_section"
        return md_chunks

    # 3. Fallback uniforme
    uniform = split_text_uniformly(text)
    return [
        {"text": chunk, "doctype": "generic", "chunk_index": i}
        for i, chunk in enumerate(uniform)
    ]


def chunk_by_type(filepath: Path, text: str, base_document_id: str) -> list[dict[str, Any]]:
    """
    Routeur de chunking intelligent : 1 concept = 1 chunk quand possible.
    """
    ext = filepath.suffix.lower()

    try:
        if ext == ".tsv":
            return chunk_glossary_table(text, delimiter="\t")
        if ext == ".csv":
            return chunk_glossary_table(text, delimiter=",")
        if ext in (".xlsx", ".xls"):
            return chunk_excel(filepath)
        if ext in (".rdf", ".ttl", ".jsonld", ".n3", ".nt", ".trig", ".owl"):
            return chunk_rdf(filepath, text, base_document_id)
        if ext == ".json":
            structured = chunk_json_structured(text)
            return structured if structured else [{"text": chunk, "doctype": "generic", "chunk_index": i} for i, chunk in enumerate(split_text_uniformly(text))]
        if ext in (".yaml", ".yml"):
            structured = chunk_yaml_structured(text)
            return structured if structured else [{"text": chunk, "doctype": "generic", "chunk_index": i} for i, chunk in enumerate(split_text_uniformly(text))]
        if ext in (".html", ".htm", ".xhtml", ".xml"):
            structured = chunk_html_glossary(text)
            return structured if structured else [{"text": chunk, "doctype": "generic", "chunk_index": i} for i, chunk in enumerate(split_text_uniformly(text))]
        if ext == ".pdf":
            return chunk_pdf_by_sections(text)
        if ext in (".txt", ".md", ".markdown", ".rst", ".adoc"):
            return chunk_text_structured(text)
        if ext == ".rtf":
            return chunk_text_structured(text)
    except Exception as e:
        print(f"[!] Structured chunking failed for {filepath.name}: {e}. Falling back to uniform chunking.")

    return [{"text": chunk, "doctype": "generic", "chunk_index": i} for i, chunk in enumerate(split_text_uniformly(text))]


def read_pdf_document(filepath: Path) -> tuple[str, str]:
    print(f"[~] Converting PDF to Markdown via pymupdf4llm: {filepath.name}")
    md_text = pymupdf4llm.to_markdown(str(filepath))
    return md_text, "pdf-to-markdown"


def read_rtf_document(filepath: Path) -> tuple[str, str]:
    print(f"[~] Converting RTF to text: {filepath.name}")
    raw = filepath.read_bytes()
    encoding = detect_bom_encoding(raw)
    text = raw.decode(encoding or "utf-8", errors="ignore")
    if rtf_to_text is not None:
        text = rtf_to_text(text)
    return text, "rtf-to-text"


# ─── Optimisation du contenu ─────────────────────────────────────────────────

def optimize_json_preserving_standards(data: Any) -> str:
    def transform(obj: Any) -> Any:
        if isinstance(obj, dict):
            out = {}
            for k, v in obj.items():
                if isinstance(v, str):
                    out[k] = v.strip()
                else:
                    out[k] = transform(v)
            return out
        if isinstance(obj, list):
            return [transform(item) for item in obj]
        if isinstance(obj, str):
            return obj.strip()
        return obj

    optimized = transform(data)
    return json.dumps(optimized, separators=(",", ":"), ensure_ascii=False)


def optimize_xml_preserving_standards(xml_content: str) -> str:
    try:
        import xml.etree.ElementTree as ET

        root = ET.fromstring(xml_content)

        def strip_whitespace(elem):
            if elem.text is not None:
                elem.text = elem.text.strip() or None
            if elem.tail is not None:
                elem.tail = elem.tail.strip() or None
            for child in list(elem):
                strip_whitespace(child)

        strip_whitespace(root)
        return ET.tostring(root, encoding="unicode", method="xml")
    except Exception:
        compact = re.sub(r">\s+<", "><", xml_content)
        return compact.strip()


def optimize_rdflike_text(content: str) -> str:
    lines = []
    previous_blank = False

    for line in content.splitlines():
        stripped = line.strip()

        if not stripped:
            if not previous_blank:
                lines.append("")
                previous_blank = True
            continue

        previous_blank = False

        if stripped.startswith("@"):
            lines.append(stripped)
            continue

        compressed = re.sub(r"\s+", " ", stripped)
        lines.append(compressed)

    return "\n".join(lines)


def optimize_yaml_content(text: str) -> str:
    try:
        import yaml

        data = yaml.safe_load(text)
        return json.dumps(data, separators=(",", ":"), ensure_ascii=False)
    except Exception:
        lines = []
        for line in text.splitlines():
            if not line.strip():
                continue
            lines.append(line.rstrip())
        return "\n".join(lines)


def optimize_csv_content(text: str) -> str:
    try:
        reader = csv.reader(StringIO(text))
        output = StringIO()
        writer = csv.writer(output, lineterminator="\n")

        for row in reader:
            if not row:
                continue
            cleaned = [cell.strip() for cell in row]
            if not any(cleaned):
                continue
            writer.writerow(cleaned)

        return output.getvalue().strip()
    except Exception:
        lines = []
        for line in text.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            lines.append(",".join(part.strip() for part in stripped.split(",")))
        return "\n".join(lines)


def optimize_html_content(text: str) -> str:
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    text = re.sub(r">\s+<", "><", text)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()


def optimize_markdown_content(text: str) -> str:
    lines = []
    previous_blank = False

    for line in text.splitlines():
        stripped = line.rstrip()

        if not stripped.strip():
            if not previous_blank:
                lines.append("")
                previous_blank = True
            continue

        lines.append(stripped)
        previous_blank = False

    return "\n".join(lines).strip()


def optimize_generic_text(text: str) -> str:
    lines = []
    previous_blank = False

    for line in text.splitlines():
        stripped = line.rstrip()

        if not stripped.strip():
            if not previous_blank:
                lines.append("")
                previous_blank = True
            continue

        lines.append(re.sub(r"\s+", " ", stripped))
        previous_blank = False

    return "\n".join(lines).strip()


def optimize_document_content(filepath: Path, text: str) -> tuple[str, bool]:
    ext = filepath.suffix.lower()

    try:
        if ext == ".json":
            data = json.loads(text)
            optimized = optimize_json_preserving_standards(data)
            return optimized, optimized != text

        if ext in (".xml", ".xmi", ".xsd"):
            optimized = optimize_xml_preserving_standards(text)
            return optimized, optimized != text

        if ext in (".ttl", ".rdf", ".owl", ".n3", ".nt", ".trig"):
            optimized = optimize_rdflike_text(text)
            return optimized, optimized != text

        if ext in (".yaml", ".yml"):
            optimized = optimize_yaml_content(text)
            return optimized, optimized != text

        if ext == ".csv":
            optimized = optimize_csv_content(text)
            return optimized, optimized != text

        if ext in (".html", ".htm", ".xhtml"):
            optimized = optimize_html_content(text)
            return optimized, optimized != text

        if ext in (".md", ".markdown", ".rst", ".adoc", ".txt", ".pdf"):
            optimized = optimize_markdown_content(text)
            return optimized, optimized != text

        optimized = optimize_generic_text(text)
        if len(optimized) < len(text) * 0.95:
            return optimized, True

        return text, False

    except Exception as e:
        print(f"[!] Optimization failed for {filepath.name}: {e}")
        return text, False


# ─── Utilitaires d'indexation ────────────────────────────────────────────────

def generate_document_id(filepath: Path) -> str:
    base = str(filepath.resolve())
    return sha256(base.encode("utf-8")).hexdigest()[:24]


def generate_chunk_stable_id(document_id: str, chunk_index: int) -> int:
    base = f"{document_id}:{chunk_index}"
    hash_bytes = sha256(base.encode("utf-8")).digest()
    return int.from_bytes(hash_bytes[:8], byteorder="big") & 0x7FFFFFFFFFFFFFFF


def split_text_uniformly(
    text: str,
    chunk_size: int = CHUNK_SIZE,
    overlap: int = CHUNK_OVERLAP,
) -> list[str]:
    text = text.strip()
    if not text:
        return []

    if chunk_size <= 0:
        raise ValueError("chunk_size must be > 0")
    if overlap < 0:
        raise ValueError("overlap must be >= 0")
    if overlap >= chunk_size:
        overlap = max(0, chunk_size // 10)

    if len(text) <= chunk_size:
        return [text]

    chunks = []
    step = chunk_size - overlap
    start = 0
    n = len(text)

    while start < n:
        end = min(start + chunk_size, n)

        if end < n:
            window_start = max(start, end - 300)
            newline_pos = text.rfind("\n", window_start, end)
            space_pos = text.rfind(" ", window_start, end)
            best_cut = max(newline_pos, space_pos)
            if best_cut > start + (chunk_size // 2):
                end = best_cut

        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)

        if end >= n:
            break

        start = max(start + step, end - overlap)

    return chunks


# ─── Gestion de la collection Qdrant ────────────────────────────────────────

def setup_collection(capabilities: dict[str, Any]) -> bool:
    collections = client.get_collections().collections
    collection_exists = any(c.name == COLLECTION for c in collections)

    if collection_exists:
        print(f"[!] Collection '{COLLECTION}' already exists.")
        choice = input("Delete and recreate? (y/n): ").strip().lower()
        if choice == "y":
            client.delete_collection(collection_name=COLLECTION)
            print(f"[✓] Deleted collection '{COLLECTION}'")
        else:
            print(f"[~] Using existing collection '{COLLECTION}'")
            return False

    if capabilities["has_dense"] and capabilities["has_sparse"]:
        client.create_collection(
            collection_name=COLLECTION,
            vectors_config={
                DENSE_VECTOR_NAME: VectorParams(
                    size=capabilities["dense_dim"],
                    distance=Distance.COSINE,
                ),
            },
            sparse_vectors_config={
                SPARSE_VECTOR_NAME: SparseVectorParams(),
            },
        )
        print(
            f"[✓] Created hybrid collection '{COLLECTION}' "
            f"(dense:{capabilities['dense_dim']}, sparse:yes)"
        )

    elif capabilities["has_dense"]:
        client.create_collection(
            collection_name=COLLECTION,
            vectors_config=VectorParams(
                size=capabilities["dense_dim"],
                distance=Distance.COSINE,
            ),
        )
        print(
            f"[✓] Created dense-only collection '{COLLECTION}' "
            f"(dense:{capabilities['dense_dim']})"
        )

    elif capabilities["has_sparse"]:
        client.create_collection(
            collection_name=COLLECTION,
            vectors_config={},
            sparse_vectors_config={
                SPARSE_VECTOR_NAME: SparseVectorParams(),
            },
        )
        print(f"[✓] Created sparse-only collection '{COLLECTION}'")

    else:
        raise ValueError("Model has neither dense nor sparse capability")

    print("[~] Creating payload index for 'tags'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="tags",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'tags' created.")

    print("[~] Creating payload index for 'document_id'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="document_id",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'document_id' created.")

    print("[~] Creating payload index for 'doctype'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="doctype",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'doctype' created.")

    print("[~] Creating payload index for 'term'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="term",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'term' created.")

    print("[~] Creating payload index for 'article'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="article",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'article' created.")

    print("[~] Creating payload index for 'concept_uri'...")
    client.create_payload_index(
        collection_name=COLLECTION,
        field_name="concept_uri",
        field_schema=PayloadSchemaType.KEYWORD,
    )
    print("[✓] Payload index for 'concept_uri' created.")

    return True


def get_existing_ids():
    try:
        existing_ids = set()
        next_offset = None

        while True:
            result = client.scroll(
                collection_name=COLLECTION,
                limit=10000,
                offset=next_offset,
                with_payload=False,
                with_vectors=False,
            )
            points, next_offset = result
            existing_ids.update(point.id for point in points)

            if next_offset is None:
                break

        print(f"[~] Found {len(existing_ids)} existing chunks")
        return existing_ids
    except Exception as e:
        print(f"[!] Failed to retrieve existing IDs: {e}")
        return set()


# ─── Encodage et construction des points ────────────────────────────────────

def _lexical_to_sparse_vector(lexical_weights: dict[Any, Any]) -> SparseVector:
    indices = [int(k) for k in lexical_weights.keys()]
    values = [float(v) for v in lexical_weights.values()]
    return SparseVector(indices=indices, values=values)


def encode_batch(texts: list[str], capabilities: dict[str, Any]) -> list[dict[str, Any]]:
    outputs = []

    if capabilities["has_dense"] and capabilities["has_sparse"]:
        result = model.encode(
            texts,
            return_dense=True,
            return_sparse=True,
            return_colbert_vecs=False,
        )

        dense_vecs = result.get("dense_vecs", [])
        lexical_weights = result.get("lexical_weights", [])

        for i in range(len(texts)):
            item = {}
            if i < len(dense_vecs):
                dense = dense_vecs[i]
                item["dense"] = dense.tolist() if hasattr(dense, "tolist") else list(dense)
            if i < len(lexical_weights):
                item["sparse"] = _lexical_to_sparse_vector(lexical_weights[i])
            outputs.append(item)

        return outputs

    if capabilities["has_dense"]:
        dense_vecs = model.encode(texts)
        for dense in dense_vecs:
            outputs.append(
                {"dense": dense.tolist() if hasattr(dense, "tolist") else list(dense)}
            )
        return outputs

    if capabilities["has_sparse"]:
        result = model.encode(
            texts,
            return_dense=False,
            return_sparse=True,
            return_colbert_vecs=False,
        )
        lexical_weights = result.get("lexical_weights", [])
        for lw in lexical_weights:
            outputs.append({"sparse": _lexical_to_sparse_vector(lw)})
        return outputs

    raise ValueError("No supported vector output available from model")


def build_point(
    docid: int,
    text: str,
    filename: str,
    encodingused: str,
    vectors: dict[str, Any],
    capabilities: dict[str, Any],
    payloadextra: dict[str, Any] | None = None,
) -> PointStruct:
    payload = {
        "text": text,
        "filename": filename,
        "encoding": encodingused,
    }
    if payloadextra:
        payload.update(payloadextra)

    if capabilities["has_dense"] and capabilities["has_sparse"]:
        return PointStruct(
            id=docid,
            vector={
                DENSE_VECTOR_NAME: vectors["dense"],
                SPARSE_VECTOR_NAME: vectors["sparse"],
            },
            payload=payload,
        )

    if capabilities["has_dense"]:
        return PointStruct(
            id=docid,
            vector=vectors["dense"],
            payload=payload,
        )

    if capabilities["has_sparse"]:
        return PointStruct(
            id=docid,
            vector={SPARSE_VECTOR_NAME: vectors["sparse"]},
            payload=payload,
        )

    raise ValueError("Unable to build point: no vectors available")


def flush_batch(
    batch_docs: list[dict[str, Any]],
    points: list[PointStruct],
    capabilities: dict[str, Any],
) -> int:
    texts_to_embed = [
        f"Source : {', '.join(item['payloadextra'].get('tags', []))}\nFichier : {item['filename']}\n\n{item['text']}"
        for item in batch_docs
    ]

    encoded_vectors = encode_batch(texts_to_embed, capabilities)

    for item, vectors in zip(batch_docs, encoded_vectors):
        point = build_point(
            docid=item["docid"],
            text=item["text"],  # Le payload UI conserve le texte brut (sans l'ajout)
            filename=item["filename"],
            encodingused=item["encoding"],
            vectors=vectors,
            capabilities=capabilities,
            payloadextra=item.get("payloadextra"),
        )
        points.append(point)
        print(
            f"[~] Loaded {item['filename']} "
            f"(chunk {item['payloadextra'].get('chunk_index', 0) + 1}/"
            f"{item['payloadextra'].get('chunk_count', 1)}) "
            f"({item['encoding']})"
        )

    client.upsert(collection_name=COLLECTION, points=points)
    uploaded = len(points)
    print(f"[✓] Uploaded batch of {uploaded}")

    points.clear()
    batch_docs.clear()

    return uploaded


# ─── Indexation principale ───────────────────────────────────────────────────

def index_documents():
    print("Detecting model capabilities...")
    capabilities = cf.MODEL_CAPABILITIES

    is_fresh = setup_collection(capabilities)
    existing_ids = set() if is_fresh else get_existing_ids()

    docs_path = Path(__file__).parent / "documents"

    total_indexed = 0
    total_skipped = 0
    total_nontext = 0
    total_optimized = 0
    total_chunks = 0

    batch_docs = []
    points = []

    if not docs_path.exists():
        print(f"[!] Directory not found: {docs_path}")
        return

    print(f"[~] Scanning documents recursively in {docs_path}")

    def pushdoc(
        filename: str,
        content: str,
        encodingused: str,
        payloadextra: dict[str, Any] | None = None,
    ):
        nonlocal total_skipped

        payloadextra = payloadextra or {}
        document_id = str(payloadextra.get("document_id"))
        chunk_index = int(payloadextra.get("chunk_index", 0))
        docid = generate_chunk_stable_id(document_id, chunk_index)

        if not is_fresh and docid in existing_ids:
            print(f"[~] Skipped duplicate: {filename} chunk={chunk_index}")
            total_skipped += 1
            return

        batch_docs.append(
            {
                "filename": filename,
                "text": content,
                "encoding": encodingused,
                "docid": docid,
                "payloadextra": payloadextra,
            }
        )

    for filepath in docs_path.rglob("*"):
        if not filepath.is_file():
            continue

        relative_path = filepath.relative_to(docs_path)
        tags = list(relative_path.parent.parts)

        ext = filepath.suffix.lower()

        try:
            file_base64 = None
            try:
                with open(filepath, "rb") as f:
                    file_base64 = base64.b64encode(f.read()).decode("utf-8")
            except Exception as e:
                print(f"[!] Warning: Could not base64 encode {filepath.name}: {e}")

            if ext == ".pdf":
                text, encodingused = read_pdf_document(filepath)
            elif ext == ".rtf":
                text, encodingused = read_rtf_document(filepath)
            else:
                text, encodingused = read_text_document(filepath)

            optimized_text, was_optimized = optimize_document_content(filepath, text)

            if was_optimized:
                total_optimized += 1
                origin_len = len(text)
                optimized_len = len(optimized_text)
                reduction = ((1 - (optimized_len / origin_len)) * 100) if origin_len else 0
                print(
                    f"[~] Optimized {filepath.name}: "
                    f"{origin_len} → {optimized_len} chars "
                    f"({reduction:.1f}% reduction)"
                )

            document_id = generate_document_id(filepath)
            structured_chunks = chunk_by_type(filepath, optimized_text, document_id)

            if not structured_chunks:
                continue

            # Normaliser en listes de textes + métadonnées
            chunks = []
            for c in structured_chunks:
                if not c.get("text", "").strip():
                    continue
                chunks.append(c)

            total_chunks += len(chunks)
            print(f"[~] Chunked {filepath.name} into {len(chunks)} structured chunks (Tags: {tags})")

            doc_summary = None

            for chunk_index, chunk_meta in enumerate(chunks):
                chunk_text = chunk_meta.get("text", "").strip()
                if not chunk_text:
                    continue

                payload_extra = {
                    "doctype": chunk_meta.get("doctype", "generic"),
                    "document_id": document_id,
                    "chunk_index": chunk_index,
                    "chunk_count": len(chunks),
                    "is_child_chunk": True,
                    "source_path": str(filepath.resolve()),
                    "source_extension": filepath.suffix.lower(),
                    "document_name": filepath.stem,
                    "tags": tags,
                    # Métadonnées optionnelles
                    "term": chunk_meta.get("term", "") or "",
                    "definition": chunk_meta.get("definition", "") or "",
                    "synonyms": chunk_meta.get("synonyms", []) or [],
                    "article": chunk_meta.get("article", "") or "",
                    "cited_articles": chunk_meta.get("cited_articles", []) or [],
                    "concept_uri": chunk_meta.get("concept_uri", "") or "",
                    "broader": chunk_meta.get("broader", []) or [],
                    "narrower": chunk_meta.get("narrower", []) or [],
                    "related": chunk_meta.get("related", []) or [],
                    "section_title": chunk_meta.get("section_title", "") or "",
                    "sheet": chunk_meta.get("sheet", "") or "",
                    "parent_key": chunk_meta.get("parent_key", "") or "",
                }

                # Chunk 0 : base64
                if chunk_index == 0:
                    if file_base64:
                        payload_extra["file_base64"] = file_base64

                pushdoc(
                    filename=filepath.name,
                    content=chunk_text,
                    encodingused=encodingused,
                    payloadextra=payload_extra,
                )

                if len(batch_docs) >= BATCH_SIZE:
                    try:
                        total_indexed += flush_batch(batch_docs, points, capabilities)
                    except Exception as e:
                        print(f"[!] Batch upload failed: {e}")
                        batch_docs.clear()
                        points.clear()

        except ValueError as e:
            print(f"[!] Skipped non-text file {filepath.name}: {e}")
            total_nontext += 1
        except Exception as e:
            print(f"[!] Failed to read {filepath.name}: {e}")

    if batch_docs:
        try:
            total_indexed += flush_batch(batch_docs, points, capabilities)
        except Exception as e:
            print(f"[!] Final batch upload failed: {e}")

    print("=" * 50)
    print("Indexing complete!")
    print(f"[✓] Chunks indexed: {total_indexed}")
    print(f"[✓] Documents optimized: {total_optimized}")
    print(f"[✓] Total chunks created: {total_chunks}")
    if total_skipped > 0:
        print(f"[~] Chunks skipped (duplicates): {total_skipped}")
    if total_nontext > 0:
        print(f"[!] Files skipped (non-text): {total_nontext}")
    print("=" * 50)


if __name__ == "__main__":
    index_documents()
